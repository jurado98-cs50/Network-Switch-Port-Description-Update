import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
from netmiko import ConnectHandler
from openpyxl import load_workbook
import threading

def show_license_on_start():
    license_text = """
MIT License

Copyright (c) 2025 Daniel Jurado

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in
all copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""
    messagebox.showinfo("License Agreement", license_text)

# --- Functions ---
def browse_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    excel_path.set(file_path)

def update_descriptions_thread():
    # Disable Run button
    run_button.config(state="disabled")
    try:
        update_descriptions()
    finally:
        run_button.config(state="normal")

def update_descriptions():
    file_path = excel_path.get()
    if not file_path:
        messagebox.showerror("Error", "Please select an Excel file.")
        return

    # Load Excel
    wb = load_workbook(file_path)
    sheet = wb.active

    # Column mapping
    description_col = 1  # Column A
    interface_col = 2    # Column B
    status_col = 3       # Column C

    # Get switch login info
    ip = ip_entry.get().strip()
    username = user_entry.get().strip()
    password = pass_entry.get().strip()

    if not ip or not username or not password:
        messagebox.showerror("Error", "Please enter switch IP, username, and password.")
        return

    # Clear output box
    output_box.delete("1.0", tk.END)

    # --- Try SSH first, then Telnet ---
    connection = None
    try:
        device = {
            "device_type": "cisco_ios",
            "ip": ip,
            "username": username,
            "password": password,
            "fast_cli": False,
        }
        connection = ConnectHandler(**device)
        connection_type = "SSH"
    except Exception:
        try:
            device = {
                "device_type": "cisco_ios_telnet",
                "ip": ip,
                "username": username,
                "password": password,
                "fast_cli": False,
            }
            connection = ConnectHandler(**device)
            connection_type = "Telnet"
        except Exception as e:
            messagebox.showerror("Connection Failed", f"SSH/Telnet failed: {e}")
            return

    try:
        # Collect valid rows first
        rows = []
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            description = row[description_col - 1] if len(row) >= description_col else None
            interface = row[interface_col - 1] if len(row) >= interface_col else None
            if interface and description:
                rows.append((idx, interface, description))

        total = len(rows)
        attempted = 0

        if total == 0:
            messagebox.showinfo("No Data", "No valid interface/description entries found.")
            return

        # Setup progress bar
        progress_bar["value"] = 0
        progress_bar["maximum"] = total
        progress_var.set(f"0/{total} interfaces updated")

        # Loop through interfaces
        for row_num, interface, description in rows:
            try:
                commands = [
                    f"interface {interface}",
                    f"description {description}",
                ]
                connection.send_config_set(commands, delay_factor=2)

                output_box.insert(tk.END, f"✅ Updated {interface} -> {description}\n")
                sheet.cell(row=row_num, column=status_col, value="Success")
            except Exception as e:
                output_box.insert(tk.END, f"❌ Failed {interface}: {e}\n")
                sheet.cell(row=row_num, column=status_col, value=f"Failure: {e}")

            attempted += 1
            progress_bar["value"] = attempted
            progress_var.set(f"{attempted}/{total} interfaces updated")
            output_box.see(tk.END)
            root.update_idletasks()

        # Save configuration on switch
        try:
            connection.save_config()
        except Exception:
            pass  # some telnet devices might not support save_config

        # Save Excel file
        wb.save(file_path)

        messagebox.showinfo("Completed", f"Processed {attempted}/{total} interfaces via {connection_type}!\nResults saved in Excel.")

    except Exception as e:
        messagebox.showerror("Error", str(e))
    finally:
        if connection:
            connection.disconnect()

# --- Tkinter GUI ---
root = tk.Tk()
root.after(100, show_license_on_start)  # 🔹 Show MIT License on startup
root.title("Cisco Interface Description Updater")

# Switch credentials
tk.Label(root, text="Switch IP:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
ip_entry = tk.Entry(root, width=30)
ip_entry.grid(row=0, column=1, padx=5, pady=5)

tk.Label(root, text="Username:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
user_entry = tk.Entry(root, width=30)
user_entry.grid(row=1, column=1, padx=5, pady=5)

tk.Label(root, text="Password:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
pass_entry = tk.Entry(root, width=30, show="*")
pass_entry.grid(row=2, column=1, padx=5, pady=5)

# Excel file selection
tk.Label(root, text="Excel File:").grid(row=3, column=0, padx=5, pady=5, sticky="e")
excel_path = tk.StringVar()
excel_entry = tk.Entry(root, textvariable=excel_path, width=30)
excel_entry.grid(row=3, column=1, padx=5, pady=5)
tk.Button(root, text="Browse", command=browse_file).grid(row=3, column=2, padx=5, pady=5, sticky="w")

# Buttons row
button_frame = tk.Frame(root)
button_frame.grid(row=4, column=0, columnspan=3, pady=10)

run_button = tk.Button(button_frame, text="Run Update", command=lambda: threading.Thread(target=update_descriptions_thread).start(), width=15)
run_button.pack(side="left", padx=10)

tk.Button(button_frame, text="Exit", command=root.destroy, width=15).pack(side="left", padx=10)

# Progress label
progress_var = tk.StringVar()
progress_label = tk.Label(root, textvariable=progress_var, font=("Arial", 10, "bold"))
progress_label.grid(row=5, column=0, columnspan=3, pady=5)

# Output box
output_box = scrolledtext.ScrolledText(root, width=70, height=15)
output_box.grid(row=6, column=0, columnspan=3, padx=5, pady=5)

# Progress bar
progress_bar = ttk.Progressbar(root, length=400, mode="determinate")
progress_bar.grid(row=7, column=0, columnspan=3, padx=10, pady=10)

root.mainloop()
